import os
import sys
import shutil
import argparse
import lxml.etree as et
from tqdm import tqdm
from openpyxl import load_workbook
from datetime import datetime
from utils import parse_header, get_list, is_valid_date_format
from config import DEFAULT_ITEMS_DIR, LOG_FILE
import logging

def process_spreadsheet(xlsx_file, base_dir, items_dir, start_row, end_row):
    """
    Processes the entire Excel spreadsheet.
    """
    try:
        wb = load_workbook(filename=xlsx_file, read_only=True)
        finished_count = 0
        for sheet in wb:
            finished_count += process_sheet(sheet, base_dir, items_dir, start_row, end_row)
        print(f"\nPrepared {finished_count} records in '{items_dir}' folder.")
    except FileNotFoundError:
        logging.error(f"Input file not found: {xlsx_file}")
        print(f"Error: Input file not found: {xlsx_file}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)

def process_sheet(sheet, base_dir, items_dir, start_row, end_row):
    if not (str(sheet['A1'].value).startswith('dc.') or str(sheet['A1'].value).startswith('filenames')):
        logging.warning(f"Skipping sheet '{sheet.title}': Does not contain a valid header row.")
        print(f"\nSkipping sheet '{sheet.title}': Does not contain a valid header row.")
        return 0

    sheet_max_row = sheet.max_row
    effective_start = max(start_row, 2)
    effective_end = min(end_row or sheet_max_row, sheet_max_row)

    if effective_start > effective_end:
        print(f"\nSkipping sheet '{sheet.title}': No rows to process in the given range.")
        return 0

    total_rows = effective_end - effective_start + 1
    print(f"\nProcessing sheet '{sheet.title}': Rows {effective_start} to {effective_end} ({total_rows} row(s)).")

    md_header = parse_header(sheet[1])
    finished_count = 0

    for row_key, row in tqdm(
        enumerate(sheet.iter_rows(min_row=effective_start, max_row=effective_end), start=effective_start),
        total=total_rows,
        desc=f"Sheet: {sheet.title}",
        unit="row",
        leave=True
    ):
        process_row(row, row_key, md_header, base_dir, items_dir)
        finished_count += 1

    return finished_count

def process_row(row, row_num, md_header, base_dir, items_dir):
    """
    Processes a single data row in a sheet.
    """
    item_dir = os.path.join(items_dir, f'item_{row_num - 1:03d}')
    os.makedirs(item_dir, exist_ok=True)
    
    data = et.Element('dublin_core', attrib={"schema": "dc"})
    fulltext_files = []

    for col_key, cell_data in enumerate(row):
        if col_key not in md_header:
            logging.warning(f"Item item_{row_num - 1:03d} (Excel Row {row_num}): Missing header for column index {col_key}. Skipping cell.")
            continue

        element_info = md_header[col_key]
        element = element_info['element']
        qualifier = element_info['qualifier']
        delimit = element_info['delimit']
        raw_content = cell_data.value

        if not element:
            continue

        if element == 'filenames':
            if raw_content:
                fulltext_files.extend(get_list(str(raw_content), delimit))
        elif element == 'foldername':
            if raw_content:
                folder_names = get_list(str(raw_content), delimit)
                for folder_name in folder_names:
                    if not base_dir:
                        logging.warning(f"Item item_{row_num - 1:03d} (Excel Row {row_num}): --base_dir not specified. Cannot process folder '{folder_name}'.")
                        continue
                    actual_folder_path = os.path.join(base_dir, folder_name)
                    if os.path.isdir(actual_folder_path):
                        for item_in_folder in os.listdir(actual_folder_path):
                            if os.path.isfile(os.path.join(actual_folder_path, item_in_folder)):
                                fulltext_files.append(os.path.join(folder_name, item_in_folder))
                    else:
                        logging.warning(f"Item item_{row_num - 1:03d} (Excel Row {row_num}): Folder '{folder_name}' not found.")
        elif raw_content is not None:
            process_metadata(data, element, qualifier, delimit, raw_content, row_num, md_header, col_key)

    write_dublin_core(data, item_dir)
    copy_fulltext_files(fulltext_files, base_dir, item_dir, row_num)

def process_metadata(data, element, qualifier, delimit, raw_content, row_num, md_header, col_key):
    """
    Processes a single metadata field.
    """
    content = raw_content

    if content is None or str(content).strip() == "":
        return
    
    if element == 'date':
        if isinstance(content, datetime):
            # If it's already a datetime object, format it
            content = content.strftime('%Y-%m-%d')
        else:
            # Handle string dates in various formats
            content_str = str(content).strip()
            if content_str:
                # Validate and normalize date formats
                if is_valid_date_format(content_str):
                    content = content_str
                else:
                    logging.warning(f"Item item_{row_num - 1:03d} (Excel Row {row_num}), Column '{md_header[col_key]['element']}.{md_header[col_key]['qualifier']}': Content '{content_str}' is not a valid date format (expected YYYY, YYYY-MM, or YYYY-MM-DD).")
                    content = content_str  # Keep original content even if invalid
            else:
                logging.warning(f"Item item_{row_num - 1:03d} (Excel Row {row_num}), Column '{md_header[col_key]['element']}.{md_header[col_key]['qualifier']}': Empty date content.")

    values = get_list(str(content), delimit)
    for value in values:
        dcvalue = et.Element('dcvalue', attrib={"element": element, "qualifier": qualifier})
        dcvalue.text = value
        data.append(dcvalue)

def write_dublin_core(data, item_dir):
    """
    Writes the dublin_core.xml file.
    """
    xml_file = os.path.join(item_dir, 'dublin_core.xml')
    xmldata = et.tostring(data, encoding='UTF-8', method='xml', xml_declaration=True, pretty_print=True)
    with open(xml_file, "wb") as item_dc:
        item_dc.write(xmldata)

def copy_fulltext_files(fulltext_files, base_dir, item_dir, row_num):
    """
    Copies full-text files to the item directory and creates the 'contents' file.
    """
    if not fulltext_files:
        return

    if not base_dir:
        logging.error(f"Item item_{row_num - 1:03d} (Excel Row {row_num}): --base_dir is required for full-text files.")
        print(f"Error: --base_dir is required for full-text files. See {LOG_FILE} for details.")
        sys.exit(1)

    contents_file_path = os.path.join(item_dir, 'contents')
    with open(contents_file_path, "w") as contents_file:
        for f in fulltext_files:
            fpath = os.path.join(base_dir, f)
            if os.path.isfile(fpath):
                shutil.copy(fpath, item_dir)
                contents_file.write(f"{os.path.basename(f)}\tbundle:ORIGINAL\n")
            else:
                logging.warning(f"Item item_{row_num - 1:03d} (Excel Row {row_num}): Missing file '{f}'.")

def main():
    """
    Main function to parse arguments and initiate processing.
    """
    parser = argparse.ArgumentParser(description="Convert Excel spreadsheet into DSpace Simple Archive Format (SAF)")
    parser.add_argument("-f", "--input_file", dest="xlsx_file", required=True, help="Input XLSX file")
    parser.add_argument("-b", "--base_dir", dest="base_dir", help="Base directory for full-text files")
    parser.add_argument("-s", "--start", type=int, default=2, help="Start processing from this Excel row (default: 2)")
    parser.add_argument("-e", "--end", type=int, help="Stop processing at this Excel row (inclusive)")
    parser.add_argument("-d", "--items_dir", dest="items_dir", default=DEFAULT_ITEMS_DIR, help=f"Output directory for converted files (default: {DEFAULT_ITEMS_DIR})")
    parser.add_argument("--log_file", dest="log_file", default=LOG_FILE, help=f"Log file path (default: {LOG_FILE})")
    args = parser.parse_args()

    # Proper logging configuration
    for handler in logging.root.handlers[:]:
        logging.root.removeHandler(handler)

    logging.basicConfig(
        filename=args.log_file,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )

    os.makedirs(args.items_dir, exist_ok=True)
    
    process_spreadsheet(args.xlsx_file, args.base_dir, args.items_dir, args.start, args.end)

if __name__ == "__main__":
    main()
